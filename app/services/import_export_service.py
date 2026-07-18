"""Bulk product import/export (CSV, Excel, JSON), plus customer/inventory export."""
import csv
import io
import json

from openpyxl import Workbook, load_workbook

from app.core.middleware.error_handler import ConflictError
from app.repositories.product_repository import ProductRepository
from app.repositories.category_repository import CategoryRepository
from app.services.product_service import ProductService

PRODUCT_COLUMNS = [
    "sku", "barcode", "name", "description", "price", "cost_price",
    "tax_rate", "category_name", "unit", "current_stock", "is_active",
]


class ImportExportService:
    def __init__(self):
        self.product_repo = ProductRepository()
        self.category_repo = CategoryRepository()
        self.product_service = ProductService()

    # ---------- Export ----------

    def export_products_csv(self):
        products = self.product_repo.get_all().all()
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=PRODUCT_COLUMNS)
        writer.writeheader()
        for p in products:
            writer.writerow(self._product_row(p))
        return buf.getvalue().encode("utf-8")

    def export_products_xlsx(self):
        products = self.product_repo.get_all().all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(PRODUCT_COLUMNS)
        for p in products:
            row = self._product_row(p)
            ws.append([row[c] for c in PRODUCT_COLUMNS])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def export_products_json(self):
        products = self.product_repo.get_all().all()
        return json.dumps([p.to_dict() for p in products], indent=2).encode("utf-8")

    def _product_row(self, p):
        return {
            "sku": p.sku, "barcode": p.barcode or "", "name": p.name,
            "description": p.description or "", "price": float(p.price),
            "cost_price": float(p.cost_price), "tax_rate": float(p.tax_rate),
            "category_name": p.category.name if p.category else "",
            "unit": p.unit, "current_stock": p.inventory.quantity if p.inventory else 0,
            "is_active": p.is_active,
        }

    def export_customers_csv(self):
        from app.repositories.customer_repository import CustomerRepository
        customers = CustomerRepository().get_all().all()
        buf = io.StringIO()
        fieldnames = ["id", "name", "email", "phone", "address", "loyalty_points", "is_active"]
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for c in customers:
            writer.writerow({k: getattr(c, k) for k in fieldnames})
        return buf.getvalue().encode("utf-8")

    def export_inventory_csv(self):
        from app.repositories.inventory_repository import InventoryRepository
        inventories = InventoryRepository().get_all().all()
        buf = io.StringIO()
        fieldnames = ["product_sku", "product_name", "quantity", "low_stock_threshold", "is_low_stock"]
        writer = csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for inv in inventories:
            writer.writerow({
                "product_sku": inv.product.sku if inv.product else "",
                "product_name": inv.product.name if inv.product else "",
                "quantity": inv.quantity, "low_stock_threshold": inv.low_stock_threshold,
                "is_low_stock": inv.is_low_stock,
            })
        return buf.getvalue().encode("utf-8")

    # ---------- Import ----------

    def import_products_csv(self, actor_id, file_bytes):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return self._import_rows(actor_id, reader)

    def import_products_xlsx(self, actor_id, file_bytes):
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() for h in next(rows_iter)]
        rows = (dict(zip(headers, row)) for row in rows_iter)
        return self._import_rows(actor_id, rows)

    def _import_rows(self, actor_id, rows):
        created, updated, errors = 0, 0, []
        for i, row in enumerate(rows, start=2):
            try:
                sku = str(row.get("sku") or "").strip()
                if not sku:
                    errors.append({"row": i, "error": "Missing SKU"})
                    continue

                category_id = None
                category_name = (row.get("category_name") or "").strip() if row.get("category_name") else None
                if category_name:
                    category = self.category_repo.get_by_name(category_name)
                    if not category:
                        category = self.category_repo.create(name=category_name)
                    category_id = category.id

                existing = self.product_repo.get_by_sku(sku)
                data = {
                    "sku": sku,
                    "barcode": (row.get("barcode") or None) or None,
                    "name": str(row.get("name") or sku),
                    "description": row.get("description") or None,
                    "price": float(row.get("price") or 0),
                    "cost_price": float(row.get("cost_price") or 0),
                    "tax_rate": float(row.get("tax_rate") or 0),
                    "category_id": category_id,
                    "unit": row.get("unit") or "pcs",
                }

                if existing:
                    self.product_service.update_product(actor_id, existing.id, data)
                    updated += 1
                else:
                    data["initial_stock"] = int(row.get("current_stock") or 0)
                    self.product_service.create_product(actor_id, data)
                    created += 1
            except Exception as exc:
                errors.append({"row": i, "error": str(exc)})

        return {"created": created, "updated": updated, "errors": errors}
